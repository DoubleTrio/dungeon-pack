#!/usr/bin/env python3
"""
Write back an edited spawn xlsx into the original JSON file in place.

Reads the Pokemon and Items sheets and reconstructs the Spawns lists inside
the matching TeamSpawnZoneStep and ItemSpawnZoneStep ZoneSteps, fully replacing
their contents with whatever rows are in the xlsx (add, edit, delete all work).

Usage:
    python writeback_spawns.py <input.json> <input.xlsx>
"""

import json
import sys
from pathlib import Path
from copy import deepcopy

import openpyxl


# ---------------------------------------------------------------------------
# Read xlsx into row dicts
# ---------------------------------------------------------------------------

def read_sheet(wb, sheet_name: str, fields: list[str]) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        print(f"Warning: sheet '{sheet_name}' not found in xlsx.", file=sys.stderr)
        return []
    ws = wb[sheet_name]
    rows = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


POKEMON_FIELDS = ["Segment", "Floor Start", "Floor End", "Species", "Form", "Level", "Moves", "Rate"]
ITEM_FIELDS    = ["Segment", "Floor Start", "Floor End", "Category", "Item ID", "Cursed", "Amount", "Rate"]


# ---------------------------------------------------------------------------
# Row -> JSON serialization
# ---------------------------------------------------------------------------

def to_floor_range(floor_start, floor_end) -> dict:
    """Convert 1-indexed display values back to 0-indexed non-inclusive JSON range."""
    fs = int(floor_start)
    fe = int(floor_end)
    return {"Min": fs - 1, "Max": fe}   # start: -1 for 0-index; end: kept as-is (non-inclusive)


def parse_level(level_str) -> tuple[int, int]:
    s = str(level_str).strip()
    # Handle floats like "4.0" that openpyxl returns for numeric cells
    if s.endswith(".0"):
        s = s[:-2]
    if "-" in s:
        parts = s.split("-", 1)
        return int(parts[0]), int(parts[1])
    v = int(s)
    return v, v


def mandatory_spawn_features(num_moves: int) -> list:
    """The three SpawnFeatures every mob entry must have."""
    return [
        {
            "$type": "PMDC.LevelGen.MobSpawnMovesOff, PMDC",
            "StartAt": num_moves,
            "Remove": False
        },
        {
            "$type": "PMDC.LevelGen.MobSpawnWeak, PMDC"
        },
        {
            "$type": "PMDC.LevelGen.MobSpawnLuaTable, PMDC",
            "LuaTable": "{ EmberfrostRun = true }"
        }
    ]


def serialize_pokemon_row(row: dict, template: dict | None) -> dict:
    """
    Build a full Spawn entry dict from a Pokemon sheet row.
    Uses template to preserve Tactic, SpawnConditions and Role.
    SpawnFeatures are always rebuilt from the mandatory set.
    """
    species   = str(row["Species"]).strip()
    form      = int(row["Form"]) if row["Form"] is not None else 0
    lv_min, lv_max = parse_level(row["Level"])
    moves_raw = row.get("Moves") or ""
    moves     = [m.strip() for m in str(moves_raw).split(",") if m.strip()]
    rate      = int(row["Rate"])
    floor_start = row["Floor Start"]
    floor_end   = row["Floor End"]

    # Start from a deep copy of the template to preserve Tactic, Role, etc.,
    # or build from scratch if this is a brand-new entry.
    if template:
        entry = deepcopy(template)
        inner = entry["Spawn"]["Spawn"]
    else:
        entry = {
            "Spawn": {
                "Spawn": {
                    "BaseForm": {"Species": "", "Form": 0, "Skin": "", "Gender": -1},
                    "Level": {"Min": 0, "Max": 0},
                    "SpecifiedSkills": [],
                    "Intrinsic": "",
                    "Tactic": "wander_normal",
                    "SpawnConditions": [],
                    "SpawnFeatures": []
                },
                "Role": 0
            },
            "Rate": 0,
            "Range": {"Min": 0, "Max": 0}
        }
        inner = entry["Spawn"]["Spawn"]

    inner["BaseForm"]["Species"] = species
    inner["BaseForm"]["Form"]    = form
    inner["Level"]["Min"]        = lv_min
    inner["Level"]["Max"]        = lv_max + 1  # non-inclusive upper bound
    inner["SpecifiedSkills"]     = moves
    # Always rebuild SpawnFeatures with the mandatory set
    inner["SpawnFeatures"]       = mandatory_spawn_features(len(moves))

    entry["Rate"]  = rate
    entry["Range"] = to_floor_range(floor_start, floor_end)
    return entry


def serialize_item_row(row: dict, template: dict | None) -> dict:
    """Build a full item Spawn entry dict from an Items sheet row."""
    item_id = str(row["Item ID"]).strip()
    cursed  = bool(row.get("Cursed") or False)
    if isinstance(row.get("Cursed"), str):
        cursed = row["Cursed"].strip().lower() == "true"
    amount  = int(row.get("Amount") or 0)
    rate    = int(row["Rate"])
    floor_start = row["Floor Start"]
    floor_end   = row["Floor End"]

    if template:
        entry = deepcopy(template)
        spawn = entry["Spawn"]
    else:
        entry = {
            "Spawn": {"ID": "", "Cursed": False, "HiddenValue": "", "Amount": 0, "Price": 0},
            "Rate": 0,
            "Range": {"Min": 0, "Max": 0}
        }
        spawn = entry["Spawn"]

    spawn["ID"]     = item_id
    spawn["Cursed"] = cursed
    spawn["Amount"] = amount

    entry["Rate"]  = rate
    entry["Range"] = to_floor_range(floor_start, floor_end)
    return entry


# ---------------------------------------------------------------------------
# Build lookup keys to match xlsx rows to existing JSON entries
# ---------------------------------------------------------------------------

def pokemon_key(species: str, form, seg: int) -> tuple:
    return (str(species).strip().lower(), int(form) if form is not None else 0, int(seg))


def item_key(item_id: str, category: str, seg: int) -> tuple:
    return (str(item_id).strip().lower(), str(category).strip().lower(), int(seg))


# ---------------------------------------------------------------------------
# Write back
# ---------------------------------------------------------------------------

def writeback(json_path: str, xlsx_path: str):
    # Load JSON
    with open(json_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    # Load xlsx
    wb = openpyxl.load_workbook(xlsx_path)
    pokemon_rows = read_sheet(wb, "Pokemon", POKEMON_FIELDS)
    item_rows    = read_sheet(wb, "Items",   ITEM_FIELDS)

    # Always target Segments[0] — ignore the Segment column in the xlsx
    pokemon_by_seg: dict[int, list[dict]] = {0: pokemon_rows}

    # Items grouped by category only, always targeting Segments[0]
    items_by_seg: dict[int, dict[str, list[dict]]] = {}
    for row in item_rows:
        category = str(row["Category"]).strip()
        items_by_seg.setdefault(0, {}).setdefault(category, []).append(row)

    root     = data.get("Object", data)
    segments = root.get("Segments", [])

    pokemon_written = 0
    items_written   = 0

    for seg_idx, segment in enumerate(segments):
        if seg_idx != 0:
            continue
        for step in segment.get("ZoneSteps", []):

            # --- Pokemon ---
            if step.get("$type") == "RogueEssence.LevelGen.TeamSpawnZoneStep, RogueEssence":
                xlsx_rows = pokemon_by_seg.get(seg_idx, [])

                # Build lookup from existing entries
                existing: dict[tuple, dict] = {}
                for entry in step.get("Spawns", []):
                    try:
                        inner   = entry["Spawn"]["Spawn"]
                        species = inner["BaseForm"]["Species"]
                        form    = inner["BaseForm"]["Form"]
                        key     = pokemon_key(species, form, seg_idx)
                        existing[key] = entry
                    except (KeyError, TypeError):
                        pass

                new_spawns = []
                for row in xlsx_rows:
                    key      = pokemon_key(row["Species"], row["Form"], seg_idx)
                    template = existing.get(key)
                    new_spawns.append(serialize_pokemon_row(row, template))

                step["Spawns"] = new_spawns
                pokemon_written += len(new_spawns)

            # --- Items ---
            elif step.get("$type") == "RogueEssence.LevelGen.ItemSpawnZoneStep, RogueEssence":
                seg_item_cats = items_by_seg.get(seg_idx, {})
                spawns_dict   = step.get("Spawns", {})

                # Collect all categories that exist in either JSON or xlsx
                all_categories = set(spawns_dict.keys()) | set(seg_item_cats.keys())

                for category in all_categories:
                    xlsx_cat_rows = seg_item_cats.get(category, [])

                    # Build lookup from existing entries for this category
                    existing: dict[tuple, dict] = {}
                    for entry in spawns_dict.get(category, {}).get("Spawns", []):
                        try:
                            key = item_key(entry["Spawn"]["ID"], category, seg_idx)
                            existing[key] = entry
                        except (KeyError, TypeError):
                            pass

                    new_spawns = []
                    for row in xlsx_cat_rows:
                        key      = item_key(row["Item ID"], category, seg_idx)
                        template = existing.get(key)
                        new_spawns.append(serialize_item_row(row, template))

                    # If category existed before, preserve its structure; otherwise create it
                    if category not in spawns_dict:
                        spawns_dict[category] = {"Spawns": []}
                    spawns_dict[category]["Spawns"] = new_spawns
                    items_written += len(new_spawns)

                    # Remove category entirely if xlsx has no rows for it
                    if not new_spawns and category not in seg_item_cats:
                        del spawns_dict[category]

                step["Spawns"] = spawns_dict

    # Write back in place (utf-8-sig to preserve BOM)
    with open(json_path, "w", encoding="utf-8-sig") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Done. Wrote {pokemon_written} pokemon spawns and {items_written} item spawns back to {json_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    for p in (json_path, xlsx_path):
        if not Path(p).exists():
            print(f"Error: file not found: {p}", file=sys.stderr)
            sys.exit(1)

    writeback(json_path, xlsx_path)


if __name__ == "__main__":
    main()