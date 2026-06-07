#!/usr/bin/env python3
"""
Parse Pokemon Mystery Dungeon zone JSON files and output spawn data.

Produces a single .xlsx file with two sheets:
  - Pokemon : spawns from TeamSpawnZoneStep
  - Items   : spawns from ItemSpawnZoneStep

Usage:
    python parse_spawns.py <input.json> [output.xlsx]

If no output file is given, defaults to <input_name>.xlsx
"""

import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def floor_range(range_data: dict) -> tuple:
    """Return (floor_start, floor_end) as 1-indexed, end non-inclusive adjusted."""
    floor_min = range_data.get("Min", "?")
    floor_max = range_data.get("Max", "?")
    floor_start = floor_min + 1 if isinstance(floor_min, int) else floor_min
    floor_end   = floor_max     if isinstance(floor_max,  int) else floor_max
    return floor_start, floor_end


# ---------------------------------------------------------------------------
# Pokemon spawns
# ---------------------------------------------------------------------------

def extract_pokemon_spawns(data: dict) -> list[dict]:
    rows = []
    root = data.get("Object", data)
    for seg_idx, segment in enumerate(root.get("Segments", [])):
        for step in segment.get("ZoneSteps", []):
            if step.get("$type") != "RogueEssence.LevelGen.TeamSpawnZoneStep, RogueEssence":
                continue
            for entry in step.get("Spawns", []):
                row = parse_pokemon_entry(entry, seg_idx)
                if row:
                    rows.append(row)
    return rows


def parse_pokemon_entry(entry: dict, seg_idx: int) -> dict | None:
    try:
        spawn_inner = entry.get("Spawn", {}).get("Spawn", {})
        base_form   = spawn_inner.get("BaseForm", {})
        species     = base_form.get("Species", "")
        form        = base_form.get("Form", 0)

        level_data  = spawn_inner.get("Level", {})
        lv_min, lv_max = level_data.get("Min", "?"), level_data.get("Max", "?")
        level_str   = str(lv_min) if lv_min == lv_max else f"{lv_min}-{lv_max}"

        moves       = spawn_inner.get("SpecifiedSkills", [])
        moves_str   = ", ".join(moves) if moves else ""

        floor_start, floor_end = floor_range(entry.get("Range", {}))

        return {
            "Segment":     seg_idx,
            "Floor Start": floor_start,
            "Floor End":   floor_end,
            "Species":     species,
            "Form":        form,
            "Level":       level_str,
            "Moves":       moves_str,
            "Rate":        entry.get("Rate", ""),
        }
    except Exception as e:
        print(f"Warning (pokemon): {e}", file=sys.stderr)
        return None


# ---------------------------------------------------------------------------
# Item spawns
# ---------------------------------------------------------------------------

def extract_item_spawns(data: dict) -> list[dict]:
    rows = []
    root = data.get("Object", data)
    for seg_idx, segment in enumerate(root.get("Segments", [])):
        for step in segment.get("ZoneSteps", []):
            if step.get("$type") != "RogueEssence.LevelGen.ItemSpawnZoneStep, RogueEssence":
                continue
            spawns_dict = step.get("Spawns", {})
            for category, category_data in spawns_dict.items():
                for entry in category_data.get("Spawns", []):
                    row = parse_item_entry(entry, seg_idx, category)
                    if row:
                        rows.append(row)
    return rows


def parse_item_entry(entry: dict, seg_idx: int, category: str) -> dict | None:
    try:
        spawn    = entry.get("Spawn", {})
        floor_start, floor_end = floor_range(entry.get("Range", {}))
        return {
            "Segment":     seg_idx,
            "Floor Start": floor_start,
            "Floor End":   floor_end,
            "Category":    category,
            "Item ID":     spawn.get("ID", ""),
            "Cursed":      spawn.get("Cursed", False),
            "Amount":      spawn.get("Amount", 0),
            "Rate":        entry.get("Rate", ""),
        }
    except Exception as e:
        print(f"Warning (item): {e}", file=sys.stderr)
        return None


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF")

POKEMON_FIELDS = ["Segment", "Floor Start", "Floor End", "Species", "Form", "Level", "Moves", "Rate"]
ITEM_FIELDS    = ["Segment", "Floor Start", "Floor End", "Category", "Item ID", "Cursed", "Amount", "Rate"]


def write_sheet(ws, fieldnames: list[str], rows: list[dict]):
    # Header row
    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def write_xlsx(pokemon_rows: list[dict], item_rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    ws_pokemon = wb.active
    ws_pokemon.title = "Pokemon"
    write_sheet(ws_pokemon, POKEMON_FIELDS, pokemon_rows)

    ws_items = wb.add_sheet("Items") if hasattr(wb, "add_sheet") else wb.create_sheet("Items")
    write_sheet(ws_items, ITEM_FIELDS, item_rows)

    wb.save(output_path)
    print(f"Written to {output_path}  ({len(pokemon_rows)} pokemon rows, {len(item_rows)} item rows)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_file(json_path: str) -> tuple[list[dict], list[dict]]:
    with open(json_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    return extract_pokemon_spawns(data), extract_item_spawns(data)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) >= 3 else Path(json_path).stem + ".xlsx"

    if not Path(json_path).exists():
        print(f"Error: file not found: {json_path}", file=sys.stderr)
        sys.exit(1)

    pokemon_rows, item_rows = parse_file(json_path)
    write_xlsx(pokemon_rows, item_rows, output_path)


if __name__ == "__main__":
    main()